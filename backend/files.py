import asyncio
import os
import json
import re
import secrets
import shutil
import threading
import time
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import FileResponse, HTMLResponse, PlainTextResponse
from pydantic import BaseModel
from .auth import require_token, require_token_query
from .settings import ROOT, atomic_write_text, env_int

router = APIRouter(prefix="/api/files", tags=["files"])

# ============================================================
# Trash / dustbin — soft delete moves to <ROOT>/.muselab-dustbin/ instead
# of unlink. Restore + permanent-purge are separate endpoints. The dir is
# always excluded from file tree listings, search, and grep (it has its
# own dedicated UI surface in the frontend).
#
# Layout per deletion:
#   <ROOT>/.muselab-dustbin/<trash_id>.json   ← manifest (original path,
#                                                deletion time, kind, size)
#   <ROOT>/.muselab-dustbin/<trash_id>        ← payload (file OR dir, the
#                                                inode rename'd in place)
# trash_id = "<unix_ts>_<8-hex>" — sortable, collision-resistant, opaque
# to the client.
# ============================================================
TRASH_DIR_NAME = ".muselab-dustbin"

# Serializes every "check destination then rename into place" sequence
# (upload finalize / rename / trash-restore). Each of those does an
# exists() probe followed by a rename() — two concurrent requests against
# the same destination could both pass the probe and the later rename
# silently clobbers the earlier file (TOCTOU). One process-wide lock is
# enough: the guarded section is pure metadata ops (exists + rename),
# microseconds each — slow streaming IO stays OUTSIDE the lock.
_DEST_WRITE_LOCK = threading.Lock()


def _trash_dir() -> Path:
    return ROOT / TRASH_DIR_NAME


def _guard_not_trash(target: Path) -> None:
    """Refuse write/upload/rename/copy operations that target the dustbin.

    /delete already blocks the dustbin (writes there have dedicated
    /trash/* endpoints with a different mental model), but the write
    endpoints used to let callers freely create / overwrite / rename
    files inside .muselab-dustbin, corrupting the soft-delete bookkeeping
    (orphan payloads, manifest mismatch). Apply the same guard everywhere
    for consistency."""
    trash_root = _trash_dir()
    if target == trash_root or trash_root in target.parents:
        raise HTTPException(
            status_code=400,
            detail="cannot write inside the dustbin — use /trash/* endpoints",
        )


def _ensure_trash_dir() -> Path:
    d = _trash_dir()
    d.mkdir(parents=True, exist_ok=True)
    return d


_TRASH_ID_RE = re.compile(r"^\d{1,20}_[0-9a-f]{8}$")


def _gen_trash_id() -> str:
    return f"{int(time.time())}_{secrets.token_hex(4)}"


def _valid_trash_id(tid: str) -> bool:
    """trash_id format check — `^\\d+_[0-9a-f]{8}$` (matches _gen_trash_id).

    Defense in depth: the trash_id flows from the user back through
    /trash/restore + /trash/purge endpoints, where it's used to build
    paths under <ROOT>/.muselab-dustbin/. Without validation, a payload
    like ``"../../etc/passwd"`` would resolve outside the trash dir and
    trash_purge would rmtree arbitrary directories. The auth token is
    the primary defense, but a strict format check costs nothing and
    blocks the exploit class entirely.
    """
    return bool(tid) and bool(_TRASH_ID_RE.fullmatch(tid))


def _dir_size(p: Path) -> int:
    """Sum of file sizes (best-effort; OSError on individual files skipped)."""
    total = 0
    try:
        for sub in p.rglob("*"):
            try:
                if sub.is_file():
                    total += sub.stat().st_size
            except OSError:
                continue
    except OSError:
        pass
    return total


def _move_to_trash(target: Path) -> dict:
    """Move `target` into the trash, write a manifest, return it.
    Caller is responsible for ensuring `target` exists + is inside ROOT.
    Same-filesystem rename, so atomic + cheap regardless of payload size."""
    trash = _ensure_trash_dir()
    tid = _gen_trash_id()
    payload = trash / tid
    original_rel = str(target.relative_to(ROOT))
    target.rename(payload)
    is_dir = payload.is_dir()
    try:
        size = _dir_size(payload) if is_dir else payload.stat().st_size
    except OSError:
        size = 0
    manifest = {
        "trash_id": tid,
        "original_path": original_rel,
        "original_name": target.name,
        "deleted_at": time.time(),
        "kind": "dir" if is_dir else "file",
        "size": size,
    }
    # Atomic write: a half-written manifest from a crash mid-rename
    # would orphan the payload (manifest parses as invalid JSON → item
    # filtered out of /trash/list → user permanently loses access to
    # their soft-deleted file). atomic_write_text uses tempfile + rename
    # so readers always see either the old or the new full content.
    atomic_write_text(trash / f"{tid}.json", json.dumps(manifest))
    return manifest


def _read_manifest(tid: str) -> dict | None:
    mf = _trash_dir() / f"{tid}.json"
    if not mf.exists():
        return None
    try:
        return json.loads(mf.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def _list_trash() -> list[dict]:
    """Manifests of every in-trash item, newest first. Orphans (manifest
    without payload, or vice versa) are skipped — they'd be confusing
    to surface and the user can't usefully act on them anyway."""
    d = _trash_dir()
    if not d.exists():
        return []
    items: list[dict] = []
    try:
        for mf in d.glob("*.json"):
            try:
                data = json.loads(mf.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            tid = data.get("trash_id")
            if not tid:
                continue
            payload = d / tid
            if not payload.exists():
                continue
            items.append(data)
    except OSError:
        return []
    items.sort(key=lambda x: x.get("deleted_at", 0), reverse=True)
    return items


# Auto-expire trash items older than this many days. Tunable via env so
# users on tiny SSDs can be more aggressive; 0 = never auto-purge.
# Default 30 days mirrors macOS Finder / GNOME's "permanently delete after
# 30 days" behaviour — long enough for "wait, I needed that", short enough
# that the dustbin doesn't silently eat the disk.
#
# env_int handles non-numeric input (typo / "30 days" / etc.) by falling
# back to the default + logging to stderr — a config mistake leaves the
# feature disabled with a clear reason instead of bricking backend import.
_TRASH_TTL_DAYS = env_int("MUSELAB_TRASH_TTL_DAYS", 30, min_value=0)


def auto_purge_expired_trash() -> int:
    """Purge trash items whose `deleted_at` is older than _TRASH_TTL_DAYS.
    Returns the count purged. Called once at startup (see backend/main.py)
    and ignores any per-item errors so a single corrupt manifest can't
    block the cleanup of healthy entries. Returns 0 when disabled."""
    if _TRASH_TTL_DAYS <= 0:
        return 0
    d = _trash_dir()
    if not d.exists():
        return 0
    cutoff = time.time() - (_TRASH_TTL_DAYS * 86400)
    purged = 0
    for mf in list(d.glob("*.json")):
        try:
            data = json.loads(mf.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if (data.get("deleted_at") or 0) >= cutoff:
            continue
        tid = data.get("trash_id")
        if not tid or not _valid_trash_id(tid):
            continue
        _purge_one(tid)
        purged += 1
    return purged


def _purge_one(tid: str) -> None:
    """Permanently delete one trash item (manifest + payload).
    Silent no-op if neither exists."""
    d = _trash_dir()
    payload = d / tid
    if payload.exists():
        if payload.is_dir():
            shutil.rmtree(payload, ignore_errors=True)
        else:
            try:
                payload.unlink()
            except OSError:
                pass
    mf = d / f"{tid}.json"
    if mf.exists():
        try:
            mf.unlink()
        except OSError:
            pass

# Filenames without extensions that are commonly text (Dockerfile, Makefile, etc.).
# Compared case-insensitively against the full name.
# Known-binary extensions — fast reject, don't even try to sniff.
# Everything NOT in this set + not containing NUL bytes in the sniff window
# is treated as text-previewable. This lets .tmpl / .vue.bak / random custom
# extensions all preview without us maintaining a whitelist.
BINARY_EXT = {
    # archives / packages
    ".zip", ".tar", ".gz", ".bz2", ".xz", ".7z", ".rar", ".tgz", ".tbz",
    ".whl", ".jar", ".war", ".ear", ".deb", ".rpm", ".pkg", ".dmg", ".iso",
    ".apk", ".ipa", ".xpi", ".crx",
    # images (have their own img preview)
    ".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".ico", ".tiff", ".tif",
    ".heic", ".heif", ".raw", ".psd", ".ai", ".sketch", ".fig",
    # audio / video
    ".mp3", ".m4a", ".wav", ".flac", ".ogg", ".opus", ".aac", ".wma",
    ".mp4", ".webm", ".mkv", ".avi", ".mov", ".wmv", ".flv", ".m4v",
    # binary docs (PDF has its own preview; office formats need conversion)
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".odt",
    ".ods", ".odp", ".rtf", ".epub", ".mobi",
    # executables / libs
    ".exe", ".dll", ".so", ".dylib", ".bin", ".o", ".a", ".lib", ".obj",
    ".class", ".pyc", ".pyo", ".elc", ".wasm",
    # fonts
    ".ttf", ".otf", ".woff", ".woff2", ".eot",
    # databases
    ".db", ".sqlite", ".sqlite3", ".mdb",
}
MAX_TEXT_SIZE = 2 * 1024 * 1024  # 2 MB — bigger files refuse with 413
SNIFF_BYTES = 4096                # how much we read to detect NUL bytes


def _looks_binary(p: Path) -> bool:
    """Heuristic: read up to 4 KB, presence of NUL byte → binary. Otherwise
    decode with `errors="replace"` and check how many bytes turned into the
    Unicode replacement character (U+FFFD). High ratio → binary / garbage.

    Important: we must NOT use plain `decode("utf-8")` here. The sniff window
    cuts at a fixed byte offset, which routinely splits a multi-byte UTF-8
    character (CJK chars are 3 bytes). A clean text file would then raise
    UnicodeDecodeError purely because of the chunk boundary — wrongly tagged
    binary. `errors="replace"` decodes whatever can be decoded and only the
    truly invalid bytes become U+FFFD."""
    try:
        with p.open("rb") as f:
            chunk = f.read(SNIFF_BYTES)
    except OSError:
        return True
    if b"\x00" in chunk:
        return True
    # decode with replacement; count how many chars are the replacement marker
    decoded = chunk.decode("utf-8", errors="replace")
    if not decoded:
        return False
    bad = decoded.count("�")
    # >5% replacement chars across a 4 KB window strongly suggests non-UTF-8
    # binary. A clean text file boundary-split mid-char contributes ≤1 bad char.
    return (bad / len(decoded)) > 0.05

# Files whose contents should never be served or overwritten through this API,
# regardless of extension. Matches against the basename (case-insensitive).
SENSITIVE_NAMES = {
    ".env", ".env.local", ".env.production", ".env.development",
    ".netrc", ".pgpass", ".npmrc", ".pypirc", ".dockercfg",
    ".htpasswd", ".htaccess",
    "credentials", "credentials.json", "service-account.json",
    "id_rsa", "id_dsa", "id_ecdsa", "id_ed25519",
    "authorized_keys", "known_hosts",
    # Shell / language history files — frequently contain pasted tokens,
    # one-off commands with secrets. Added when MUSELAB_ROOT=$HOME became
    # supported (2026-05-17) so a token leak doesn't expose them.
    ".bash_history", ".zsh_history", ".python_history", ".node_repl_history",
    ".sqlite_history", ".lesshst", ".viminfo", ".wget-hsts",
    ".npm-debug.log", ".yarn-error.log",
}
# Extension suffixes treated as sensitive — private keys, cert bundles, and
# `.env`-style files regardless of basename (prod.env, staging.env, etc.).
SENSITIVE_SUFFIX = {".pem", ".key", ".p12", ".pfx", ".keystore", ".jks", ".env"}


def _is_sensitive(p: Path) -> bool:
    name = p.name.lower()
    if name in SENSITIVE_NAMES:
        return True
    if name.startswith(".env."):  # .env.* variants like .env.local
        return True
    if p.suffix.lower() in SENSITIVE_SUFFIX:
        return True
    return False


def safe_resolve(rel: str, allow_sensitive: bool = False) -> Path:
    """Resolve a path relative to ROOT, blocking traversal outside ROOT and,
    by default, blocking access to credential-shaped filenames.

    Defends against:
      - `../../etc/passwd` style traversal (resolve() + ROOT-prefix check)
      - **Symlink escape**: a symlink inside ROOT pointing to /etc/shadow would
        previously slip through because `.resolve()` follows symlinks. We
        explicitly check the resolved target is still under ROOT.
      - `.env`, `id_rsa`, `*.pem` etc. (SENSITIVE_SUFFIX / SENSITIVE_NAMES)."""
    rel = (rel or "").lstrip("/")
    # NUL byte in a path raises ValueError from (ROOT / rel) and FastAPI
    # converts that to a 500 with a traceback that leaks internal module
    # paths. Reject early as 400. Same for any string that Python's path
    # layer refuses (control chars trip OS-level checks downstream).
    if "\x00" in rel:
        raise HTTPException(status_code=400, detail="invalid path")
    # First-pass resolve (follows symlinks → catches symlink escape):
    try:
        target = (ROOT / rel).resolve()
    except (ValueError, OSError):
        raise HTTPException(status_code=400, detail="invalid path") from None
    # ROOT itself must also be resolved (it might itself be a symlink target).
    root_real = ROOT.resolve()
    if root_real != target and root_real not in target.parents:
        raise HTTPException(status_code=400, detail="path escapes root")
    # Block by name regardless of whether the file already exists, so the API
    # can neither read nor write `.env` / private-key shaped paths.
    if not allow_sensitive and not target.is_dir() and _is_sensitive(target):
        raise HTTPException(status_code=403, detail="sensitive file blocked")
    return target


class Entry(BaseModel):
    name: str
    path: str  # relative to ROOT
    is_dir: bool
    size: int
    mtime: float


MAX_LIST_ENTRIES = 500  # safety cap so huge dirs (.git/objects) don't freeze the UI


@router.get("/list", dependencies=[Depends(require_token)])
def list_dir(path: str = "", show_hidden: bool = False) -> dict:
    target = safe_resolve(path)
    if not target.exists():
        raise HTTPException(status_code=404, detail="not found")
    if not target.is_dir():
        raise HTTPException(status_code=400, detail="not a directory")
    entries: list[dict] = []
    truncated = False
    # Trash dir is always hidden from the file tree (even when
    # show_hidden=true) — it has its own dedicated UI surface; mixing it
    # back into the tree would surface deleted files in a confusing
    # context. Only relevant at the root level since trash dir lives there.
    is_root_listing = (target == ROOT)
    for child in sorted(target.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower())):
        if is_root_listing and child.name == TRASH_DIR_NAME:
            continue
        if not show_hidden and child.name.startswith("."):
            continue
        if len(entries) >= MAX_LIST_ENTRIES:
            truncated = True
            break
        try:
            stat = child.stat()
        except OSError:
            continue
        entries.append(Entry(
            name=child.name,
            path=str(child.relative_to(ROOT)),
            is_dir=child.is_dir(),
            size=stat.st_size if not child.is_dir() else 0,
            mtime=stat.st_mtime,
        ).model_dump())
    return {"root": str(ROOT), "path": path, "entries": entries, "truncated": truncated}


# xlsx preview caps. Read-only mode + capped per-sheet rows/cols so a
# 1M-cell spreadsheet doesn't OOM the SSE event loop or blow up the JSON
# payload over the wire. Truncation is signaled to the FE so it can hint
# the user instead of silently dropping data.
XLSX_MAX_SHEETS = 20
XLSX_MAX_ROWS = 500
XLSX_MAX_COLS = 50
XLSX_CELL_MAX_CHARS = 500   # one obnoxious cell shouldn't blow the page


@router.get("/xlsx", dependencies=[Depends(require_token)])
def xlsx_preview(path: str) -> dict:
    """Read-only xlsx preview as structured JSON.

    Returns each sheet's first XLSX_MAX_ROWS×XLSX_MAX_COLS cells as
    strings. Formulas are NOT evaluated — `data_only=True` returns the
    cached value the spreadsheet app last wrote. If a file was created
    programmatically without ever being opened in Excel/LibreOffice,
    formula cells will be null and surface as empty strings.
    """
    target = safe_resolve(path)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="not a file")
    if target.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise HTTPException(status_code=415, detail="not an xlsx-family file")
    try:
        import openpyxl  # local import — openpyxl is only loaded on demand
    except ImportError:
        raise HTTPException(status_code=500,
                            detail="openpyxl not installed — run `uv sync`")
    try:
        wb = openpyxl.load_workbook(target, read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=422,
                            detail=f"failed to parse xlsx: {type(e).__name__}: {e}")
    try:
        sheets: list[dict] = []
        sheet_names = wb.sheetnames
        sheets_truncated = len(sheet_names) > XLSX_MAX_SHEETS
        for sheet_name in sheet_names[:XLSX_MAX_SHEETS]:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            rows_truncated = False
            cols_truncated = False
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx >= XLSX_MAX_ROWS:
                    rows_truncated = True
                    break
                cells: list[str] = []
                for c_idx, val in enumerate(row):
                    if c_idx >= XLSX_MAX_COLS:
                        cols_truncated = True
                        break
                    if val is None:
                        cells.append("")
                    else:
                        s = str(val)
                        if len(s) > XLSX_CELL_MAX_CHARS:
                            s = s[:XLSX_CELL_MAX_CHARS] + "…"
                        cells.append(s)
                rows.append(cells)
            sheets.append({
                "name": sheet_name,
                "rows": rows,
                "rows_truncated": rows_truncated,
                "cols_truncated": cols_truncated,
            })
        return {
            "path": path,
            "sheets": sheets,
            "sheets_truncated": sheets_truncated,
            "limits": {"max_rows": XLSX_MAX_ROWS, "max_cols": XLSX_MAX_COLS,
                       "max_sheets": XLSX_MAX_SHEETS},
        }
    finally:
        wb.close()


# CSV preview caps. Paginated by design — CSV files in the wild can be
# millions of rows, so we never load the whole file into memory. Each
# request returns one window; the UI calls back with offset += limit when
# the user pages forward.
CSV_DEFAULT_LIMIT = 200       # default page size
CSV_MAX_LIMIT = 1000          # hard ceiling the client can request
CSV_MAX_COLS = 50             # per-row column cap
CSV_CELL_MAX_CHARS = 500
CSV_SNIFF_BYTES = 8192        # sample size for delimiter / header detection


@router.get("/csv", dependencies=[Depends(require_token)])
def csv_preview(path: str, offset: int = 0, limit: int = CSV_DEFAULT_LIMIT) -> dict:
    """Read-only paginated CSV / TSV preview as structured JSON.

    Returns rows[offset : offset+limit] from the file, plus the sniffed
    delimiter and a `total_rows` count so the UI can show pagination.
    Header row (if csv.Sniffer flags one) is returned separately.

    Designed to never load more than a window into memory: the file is
    iterated row-by-row, skipping rows below offset and breaking once
    `limit` is filled. The trailing total-rows count is the only full
    scan, and it just discards each row.
    """
    import csv as _csv  # local import — csv is stdlib, but keep import local
                       # so import overhead stays out of every other route.
    target = safe_resolve(path)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="not a file")
    if target.suffix.lower() not in {".csv", ".tsv"}:
        raise HTTPException(status_code=415, detail="not a csv/tsv file")
    if limit < 1:
        limit = CSV_DEFAULT_LIMIT
    if limit > CSV_MAX_LIMIT:
        limit = CSV_MAX_LIMIT
    if offset < 0:
        offset = 0
    # Sniff delimiter + header from a small head sample. Defaults to
    # excel-style comma if Sniffer can't tell (e.g. one-column file).
    try:
        with target.open("r", encoding="utf-8", errors="replace", newline="") as f:
            sample = f.read(CSV_SNIFF_BYTES)
        try:
            dialect = _csv.Sniffer().sniff(sample, delimiters=",\t;|")
            has_header = _csv.Sniffer().has_header(sample)
        except _csv.Error:
            dialect = _csv.excel
            has_header = False
        # Override sniff for explicit .tsv — Sniffer sometimes guesses comma
        # on tab-separated files when the first row has no tabs.
        if target.suffix.lower() == ".tsv":
            dialect = _csv.excel_tab
    except OSError as e:
        raise HTTPException(status_code=500, detail=f"failed to read: {e}")

    header: list[str] = []
    rows: list[list[str]] = []
    cols_truncated = False
    total_rows = 0
    try:
        with target.open("r", encoding="utf-8", errors="replace", newline="") as f:
            reader = _csv.reader(f, dialect=dialect)
            # Pull header before any data offset is applied. The user paging
            # to offset=200 still wants column titles at the top of the page.
            if has_header:
                try:
                    header_row = next(reader)
                    header = [_clip_cell(c) for c in header_row[:CSV_MAX_COLS]]
                    if len(header_row) > CSV_MAX_COLS:
                        cols_truncated = True
                except StopIteration:
                    pass
            row_idx = 0
            for raw in reader:
                if row_idx < offset:
                    row_idx += 1
                    continue
                if len(rows) < limit:
                    cells = [_clip_cell(c) for c in raw[:CSV_MAX_COLS]]
                    if len(raw) > CSV_MAX_COLS:
                        cols_truncated = True
                    rows.append(cells)
                row_idx += 1
            total_rows = row_idx
    except OSError as e:
        raise HTTPException(status_code=500, detail=f"failed to read: {e}")

    return {
        "path": path,
        "header": header,
        "rows": rows,
        "offset": offset,
        "limit": limit,
        "total_rows": total_rows,
        "has_header": has_header,
        "delimiter": dialect.delimiter,
        "cols_truncated": cols_truncated,
        "limits": {"max_cols": CSV_MAX_COLS, "max_limit": CSV_MAX_LIMIT},
    }


def _clip_cell(value: str) -> str:
    """Cap a single CSV cell so one runaway value can't blow up the page."""
    s = "" if value is None else str(value)
    if len(s) > CSV_CELL_MAX_CHARS:
        s = s[:CSV_CELL_MAX_CHARS] + "…"
    return s


@router.get("/read", dependencies=[Depends(require_token)])
def read_file(path: str) -> PlainTextResponse:
    target = safe_resolve(path)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="not a file")
    suffix = target.suffix.lower()
    # Fast reject for known binary extensions.
    if suffix in BINARY_EXT:
        raise HTTPException(status_code=415, detail="binary file — not previewable as text")
    # Single stat() reused for both the size gate and the empty-file check.
    # The previous code called target.stat() twice; if the file vanished
    # between the two calls (TOCTOU) the second stat raised
    # FileNotFoundError → 500 instead of a clean 404.
    try:
        st_size = target.stat().st_size
    except OSError:
        raise HTTPException(status_code=404, detail="not a file") from None
    if st_size > MAX_TEXT_SIZE:
        raise HTTPException(status_code=413, detail="file too large for preview")
    # Empty extension + not a known text name? Sniff content. Empty files OK.
    # This is the path that picks up .tmpl, .conf.j2, .env.staging, etc.
    if st_size > 0 and _looks_binary(target):
        raise HTTPException(status_code=415, detail="binary content — not previewable as text")
    content = target.read_text(encoding="utf-8", errors="replace")
    if len(content) > MAX_TEXT_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File too large to read as text ({len(content)} bytes > {MAX_TEXT_SIZE})",
        )
    return PlainTextResponse(
        content,
        headers={
            "Content-Disposition": "inline",
            "X-Content-Type-Options": "nosniff",
        },
    )


@router.get("/stat", dependencies=[Depends(require_token)])
def stat_file(path: str) -> dict:
    """Lightweight metadata for a single path — name, size, mtime, is_dir.

    Powers the preview header's "real path + last-modified" strip: the
    frontend only knows a tab's path string, not its on-disk mtime (the
    tree-list carries mtime but a file opened via chat-link / search is
    never in the visible tree). One cheap stat() fills that gap without
    re-reading the whole file. 404 when the path is gone — same contract
    as /read, so a stale/phantom tab surfaces honestly instead of showing
    a path that no longer exists."""
    target = safe_resolve(path)
    try:
        st = target.stat()
    except OSError:
        raise HTTPException(status_code=404, detail="not found") from None
    is_dir = target.is_dir()
    return {
        "path": str(target.relative_to(ROOT)),
        "name": target.name,
        "is_dir": is_dir,
        "size": 0 if is_dir else st.st_size,
        "mtime": st.st_mtime,
    }


# Types we serve inline (images / PDF / media render natively in browser).
INLINE_OK_SUFFIX = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".ico",
                    ".pdf", ".mp4", ".webm", ".mp3", ".ogg", ".wav"}
# Types we serve inline INSIDE A SANDBOXED IFRAME (HTML / SVG can render but
# the strong CSP + sandbox attribute on the iframe blocks JS execution and
# same-origin token theft).
SANDBOXED_INLINE_SUFFIX = {".html", ".htm", ".svg"}

# Click-to-zoom bridge for HTML previews. The preview iframe is sandboxed
# (opaque origin) so the parent can't reach the framed DOM to intercept image
# clicks. Instead we inject this tiny script (only when the preview pane asks
# via ?preview=1) which forwards in-page image clicks to the parent via
# postMessage — postMessage works fine from a sandboxed origin, so this needs
# NO sandbox relaxation and leaks nothing (only the clicked image's src/alt).
# The parent (app.js) validates event.source === the preview iframe before
# opening its lightbox. Images wrapped in a link are left alone so the link
# still navigates. CSP allows it: script-src includes 'unsafe-inline'.
_PREVIEW_IMG_BRIDGE = (
    "<script>(function(){document.addEventListener('click',function(e){"
    "var t=e.target;var img=t&&t.closest?t.closest('img'):null;"
    "if(!img||img.closest('a'))return;var src=img.currentSrc||img.src;"
    "if(!src)return;e.preventDefault();"
    "try{parent.postMessage({__muselab:'preview-img',src:src,alt:img.alt||''},'*');}"
    "catch(_e){}},true);})();</script>"
)
# Cap the in-memory read used for injection. Bigger HTML (e.g. reports with
# megabytes of base64 images) falls back to streaming untouched — it just
# won't have click-to-zoom, which is an acceptable degradation.
_PREVIEW_INJECT_MAX_BYTES = 12 * 1024 * 1024


def _inject_preview_img_bridge(target: Path) -> str | None:
    """Return the HTML text with the click-to-zoom bridge injected, or None to
    signal "fall back to streaming the file untouched" (too big / not utf-8)."""
    try:
        if target.stat().st_size > _PREVIEW_INJECT_MAX_BYTES:
            return None
        html = target.read_text(encoding="utf-8")
    except (OSError, UnicodeDecodeError):
        return None
    lower = html.lower()
    idx = lower.rfind("</body>")
    if idx == -1:
        idx = lower.rfind("</html>")
    if idx == -1:
        return html + _PREVIEW_IMG_BRIDGE
    return html[:idx] + _PREVIEW_IMG_BRIDGE + html[idx:]


@router.get("/raw", dependencies=[Depends(require_token_query)])
def raw_file(path: str = Query(...), preview: bool = Query(False)):
    """Stream raw file (images, PDF, sandboxed HTML, etc.). Token via query.
    Everything outside the whitelists is forced to download as octet-stream."""
    target = safe_resolve(path)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="not a file")
    suffix = target.suffix.lower()
    # `no-cache` (NOT no-store) — let browsers cache but force a conditional
    # GET (If-None-Match / If-Modified-Since) every time. FileResponse still
    # sends ETag + Last-Modified, so unchanged files return 304 cheaply; the
    # moment mtime changes, the etag flips and the browser pulls the new
    # body. Without this, browsers happily served the disk-cached version on
    # every page reload (URLs identical) and edits never showed until users
    # hit the manual reload button — see 2026-05-18 dark-mode HTML report bug.
    base_headers = {
        "X-Content-Type-Options": "nosniff",
        "Cache-Control": "no-cache",
    }
    # RFC 5987: non-ASCII filenames must be URL-encoded in filename* attribute.
    # HTTP headers are latin-1 only; Chinese / emoji filenames break encode().
    from urllib.parse import quote
    disp_filename = f'filename="file{suffix}"; filename*=UTF-8\'\'{quote(target.name)}'

    if suffix in INLINE_OK_SUFFIX:
        return FileResponse(target, headers={
            **base_headers,
            "Content-Disposition": f"inline; {disp_filename}",
        })
    if suffix in SANDBOXED_INLINE_SUFFIX:
        # CSP relaxed enough for academic HTML reports (MathJax / KaTeX / highlight.js
        # from CDN, inline <script>window.MathJax = {...}</script> config blocks).
        # The server-side `sandbox allow-scripts` DIRECTIVE puts JS in a unique
        # opaque origin even when the file is opened TOP-LEVEL (URL pasted into
        # the address bar): scripts still run, but cannot act as our origin —
        # /api/* fetches become cross-origin (CORS-blocked), cookies/storage
        # are unavailable, so the query token can't be replayed against the API.
        # Previously only the frontend iframe's sandbox attribute provided this
        # isolation, which a top-level open silently bypassed.
        sandbox_headers = {
            **base_headers,
            "Content-Disposition": f"inline; {disp_filename}",
            "Content-Security-Policy": (
                "sandbox allow-scripts; "
                "default-src 'none'; "
                "script-src 'self' 'unsafe-inline' https:; "
                "style-src 'self' 'unsafe-inline' https:; "
                "img-src 'self' data: https:; "
                "font-src https: data:; "
                "connect-src 'self'; "
                "base-uri 'none'; form-action 'none'"
            ),
        }
        # Preview pane requests ?preview=1 for HTML so we can inject the
        # click-to-zoom bridge. SVG and the top-level/download paths never
        # get it (no preview flag) and stream untouched.
        if preview and suffix in (".html", ".htm"):
            injected = _inject_preview_img_bridge(target)
            if injected is not None:
                return HTMLResponse(content=injected, headers=sandbox_headers)
        return FileResponse(target, headers=sandbox_headers)
    # FileResponse(filename=) sets Content-Disposition itself; use our safe one.
    return FileResponse(target, media_type="application/octet-stream", headers={
        **base_headers,
        "Content-Disposition": f"attachment; {disp_filename}",
    })


@router.get("/download", dependencies=[Depends(require_token_query)])
def download_file(path: str = Query(...)) -> FileResponse:
    target = safe_resolve(path)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="not a file")
    from urllib.parse import quote
    suffix = target.suffix.lower()
    disp = f'attachment; filename="file{suffix}"; filename*=UTF-8\'\'{quote(target.name)}'
    return FileResponse(target, media_type="application/octet-stream",
                        headers={"Content-Disposition": disp})


class WriteReq(BaseModel):
    path: str
    content: str


# Upper bound on a single editor-save payload. Generous enough for real-world
# documents (a 10 MB Markdown file is ~3 million words) but stops a runaway
# script from filling the disk via this endpoint. Matches the spirit of
# MAX_TEXT_SIZE on the read path.
MAX_WRITE_BYTES = 10 * 1024 * 1024


@router.put("/write", dependencies=[Depends(require_token)])
def write_file(req: WriteReq) -> dict:
    """Overwrite a file at `path` with `content`. Atomic (tmpfile + rename),
    so a crash mid-write leaves the previous content intact instead of a
    truncated half-file. Capped at MAX_WRITE_BYTES to prevent the editor
    from accidentally serving as an unbounded ingest path."""
    target = safe_resolve(req.path)
    _guard_not_trash(target)
    if target.exists() and target.is_dir():
        raise HTTPException(status_code=400, detail="path is a directory")
    # Two-stage size gate. Each char is 1-4 UTF-8 bytes, so the upper
    # bound `len(s) * 4` is cheap (no encoding); if it already exceeds
    # the limit we reject without materializing the encoded bytes at
    # all. Only the borderline case (str length close to limit) needs
    # the precise encode. Saves ~10 MB transient RSS on a max-size
    # payload that was previously rejected anyway.
    char_len = len(req.content)
    if char_len * 4 > MAX_WRITE_BYTES:
        if char_len > MAX_WRITE_BYTES \
                or len(req.content.encode("utf-8")) > MAX_WRITE_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"content exceeds {MAX_WRITE_BYTES // (1024 * 1024)} MB limit",
            )
    target.parent.mkdir(parents=True, exist_ok=True)
    atomic_write_text(target, req.content)
    return {"ok": True, "size": target.stat().st_size}


# Default 100 MB cap per uploaded file. Override via MUSELAB_MAX_UPLOAD_MB.
MAX_UPLOAD_BYTES = env_int("MUSELAB_MAX_UPLOAD_MB", 100, min_value=1) * 1024 * 1024
# Filename extensions that are likely to be hostile or pointless to host in
# a personal archive. Block at upload (cleaner than after-the-fact cleanup).
UPLOAD_BLOCKED_SUFFIX = {
    ".exe", ".dll", ".so", ".dylib", ".scr", ".com", ".bat", ".cmd",
    ".ps1",  # PowerShell scripts — block by default; allow via .env override later
    ".msi", ".app",
}


@router.post("/upload", dependencies=[Depends(require_token)])
async def upload(path: str = Form(""), file: UploadFile = File(...)) -> dict:
    target_dir = safe_resolve(path)
    _guard_not_trash(target_dir)
    if not target_dir.exists() or not target_dir.is_dir():
        raise HTTPException(status_code=400, detail="target dir invalid")
    safe_name = Path(file.filename or "upload.bin").name
    # Path("." ).name and Path("..").name are both "" — those filenames
    # produced an empty safe_name → `target_dir / ""` == target_dir, and
    # the directory checks below raised 500 instead of a clean 400.
    if not safe_name or safe_name in {".", ".."}:
        raise HTTPException(status_code=400, detail="invalid filename")
    # Block dangerous extensions early.
    suffix = Path(safe_name).suffix.lower()
    if suffix in UPLOAD_BLOCKED_SUFFIX:
        raise HTTPException(status_code=400,
                             detail=f"upload blocked by extension: {suffix}")
    # Also block uploads with sensitive filenames (.env, id_rsa etc.).
    if _is_sensitive(Path(safe_name)):
        raise HTTPException(status_code=403,
                             detail="sensitive filename blocked")
    dest = target_dir / safe_name
    # Stream + enforce size cap. Write to a temporary file first, then
    # atomically rename to dest so a crash or size-exceeded abort never
    # leaves a partial file at the intended path.
    import uuid as _uuid
    tmp_path = dest.parent / f".~{dest.name}.{_uuid.uuid4().hex[:8]}.uploading"
    written = 0
    try:
        with tmp_path.open("wb") as f:
            while chunk := await file.read(1024 * 1024):
                written += len(chunk)
                if written > MAX_UPLOAD_BYTES:
                    f.close()
                    tmp_path.unlink(missing_ok=True)
                    raise HTTPException(
                        status_code=413,
                        detail=f"upload exceeds {MAX_UPLOAD_BYTES // (1024*1024)} MB cap",
                    )
                # Off-load the blocking disk write so a large multi-MB upload
                # doesn't stall the event loop chunk-by-chunk. (perf: RED —
                # files.py upload sync write)
                await asyncio.to_thread(f.write, chunk)
        # Overwrite protection: a same-name upload used to silently clobber the
        # existing file via rename() — no 409, no trash, no undo — which
        # contradicts /rename's 409 guard and the whole soft-delete design.
        # Move the old file to trash first (recoverable) only AFTER the new
        # upload fully streamed to tmp, so a failed/aborted upload never
        # destroys the existing file. A name colliding with a directory can't
        # be auto-resolved → 409.
        # The check + trash + rename runs as ONE unit under _DEST_WRITE_LOCK
        # in a thread (lock is sync; never hold it on the event loop). Two
        # concurrent same-name uploads previously both passed exists() and
        # the later rename clobbered the earlier file with no trash entry.
        def _finalize():
            with _DEST_WRITE_LOCK:
                trashed = None
                if dest.exists():
                    if dest.is_dir():
                        raise HTTPException(
                            status_code=409,
                            detail=f"a directory named {safe_name!r} already exists here",
                        )
                    trashed = _move_to_trash(dest)
                tmp_path.rename(dest)
                return trashed
        trashed = await asyncio.to_thread(_finalize)
    except HTTPException:
        tmp_path.unlink(missing_ok=True)
        raise
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise
    return {
        "ok": True,
        "path": str(dest.relative_to(ROOT)),
        "size": dest.stat().st_size,
        # Non-null when an existing same-name file was moved to trash so the
        # frontend can surface "replaced (old version in trash)".
        "replaced_trash_id": (trashed or {}).get("trash_id"),
    }


class DeleteReq(BaseModel):
    path: str


@router.delete("/delete", dependencies=[Depends(require_token)])
def delete(req: DeleteReq, permanent: bool = Query(default=False)) -> dict:
    """Soft delete by default: move into <ROOT>/.muselab-dustbin/. The
    previous "must be empty for dirs" guard is dropped because the
    operation is now reversible — Restore moves the payload back to its
    original path. Set ?permanent=true for hard delete (skips trash).

    Refuses to delete the trash dir itself or anything inside it via this
    route — those operations have dedicated /trash/* endpoints with a
    different mental model."""
    target = safe_resolve(req.path)
    if not target.exists():
        raise HTTPException(status_code=404, detail="not found")
    trash_root = _trash_dir()
    if target == trash_root or trash_root in target.parents:
        raise HTTPException(
            status_code=400,
            detail="cannot delete trash via /delete — use /trash/purge or /trash/empty",
        )
    if permanent:
        # ignore_errors mirrors the soft-delete _purge_one path. A
        # permission / busy-file error here previously bubbled up as a 500
        # with a traceback that leaked absolute internal paths.
        if target.is_dir():
            shutil.rmtree(target, ignore_errors=True)
        else:
            try:
                target.unlink(missing_ok=True)
            except OSError:
                pass
        return {"ok": True, "permanent": True}
    manifest = _move_to_trash(target)
    return {"ok": True, "permanent": False,
            "trash_id": manifest["trash_id"], "manifest": manifest}


# ============================================================
# Trash management endpoints
# ============================================================
@router.get("/trash/list", dependencies=[Depends(require_token)])
def trash_list() -> dict:
    """All trash items, newest first. Each item: trash_id, original_path,
    original_name, deleted_at (unix sec, float), kind ('file'|'dir'), size.

    Top-level keys also include ``total_size`` (sum of all item sizes,
    bytes) and ``ttl_days`` (auto-purge horizon — 0 means disabled) so
    the UI can surface "Trash · 142 MB · auto-purged after 30 days"
    without a separate roundtrip."""
    items = _list_trash()
    return {
        "items": items,
        "total_size": sum(int(i.get("size") or 0) for i in items),
        "ttl_days": _TRASH_TTL_DAYS,
    }


class TrashIdReq(BaseModel):
    trash_id: str


@router.post("/trash/restore", dependencies=[Depends(require_token)])
def trash_restore(req: TrashIdReq) -> dict:
    """Move the payload back to its original path. Fails 409 if that path
    is now occupied — user has to rename / clear it before restoring.
    Manifest is removed on success."""
    if not _valid_trash_id(req.trash_id):
        raise HTTPException(status_code=400, detail="invalid trash_id")
    data = _read_manifest(req.trash_id)
    if not data:
        raise HTTPException(status_code=404, detail="trash item not found")
    payload = _trash_dir() / req.trash_id
    if not payload.exists():
        raise HTTPException(status_code=404, detail="trash payload missing")
    orig_rel = data.get("original_path") or ""
    if not orig_rel:
        raise HTTPException(status_code=500, detail="manifest missing original_path")
    # Reuse safe_resolve so the same anti-traversal + sensitive-name guards
    # apply to restoration as to any other write. allow_sensitive=True
    # because the user already had this file in-place before deletion;
    # blocking restore would leave their data stranded in trash.
    orig = safe_resolve(orig_rel, allow_sensitive=True)
    # Atomic check+rename — same TOCTOU shape as /rename: a concurrent
    # write landing at `orig` between the probe and the rename would be
    # silently replaced by the restored payload.
    with _DEST_WRITE_LOCK:
        if orig.exists():
            raise HTTPException(
                status_code=409,
                detail="original path is occupied; rename or clear it first",
            )
        orig.parent.mkdir(parents=True, exist_ok=True)
        payload.rename(orig)
    mf = _trash_dir() / f"{req.trash_id}.json"
    if mf.exists():
        try:
            mf.unlink()
        except OSError:
            pass
    return {"ok": True, "restored_path": orig_rel}


@router.delete("/trash/purge", dependencies=[Depends(require_token)])
def trash_purge(req: TrashIdReq) -> dict:
    """Permanently delete one trash item. Irreversible."""
    if not _valid_trash_id(req.trash_id):
        raise HTTPException(status_code=400, detail="invalid trash_id")
    d = _trash_dir()
    if not (d / f"{req.trash_id}.json").exists() and not (d / req.trash_id).exists():
        raise HTTPException(status_code=404, detail="trash item not found")
    _purge_one(req.trash_id)
    return {"ok": True}


@router.delete("/trash/empty", dependencies=[Depends(require_token)])
def trash_empty() -> dict:
    """Permanently delete every trash item. Irreversible."""
    d = _trash_dir()
    if not d.exists():
        return {"ok": True, "purged": 0}
    count = 0
    for mf in list(d.glob("*.json")):
        tid = mf.stem
        _purge_one(tid)
        count += 1
    return {"ok": True, "purged": count}


class MkdirReq(BaseModel):
    path: str


@router.post("/mkdir", dependencies=[Depends(require_token)])
def mkdir(req: MkdirReq) -> dict:
    target = safe_resolve(req.path)
    _guard_not_trash(target)
    target.mkdir(parents=True, exist_ok=True)
    return {"ok": True, "path": str(target.relative_to(ROOT))}


class RenameReq(BaseModel):
    src: str
    dst: str   # relative to ROOT


@router.post("/rename", dependencies=[Depends(require_token)])
def rename(req: RenameReq) -> dict:
    src = safe_resolve(req.src)
    dst = safe_resolve(req.dst)
    _guard_not_trash(src)
    _guard_not_trash(dst)
    # Atomic check+rename under the destination lock — without it, two
    # concurrent renames onto the same dst both pass the exists() probe
    # and the later rename silently replaces the earlier file (TOCTOU).
    with _DEST_WRITE_LOCK:
        if not src.exists():
            raise HTTPException(status_code=404, detail="source not found")
        if dst.exists():
            raise HTTPException(status_code=409, detail="destination already exists")
        dst.parent.mkdir(parents=True, exist_ok=True)
        src.rename(dst)
    return {"ok": True, "path": str(dst.relative_to(ROOT))}


# ============================================================
# Copy as .bak — the only "copy" we expose. Frontend supports both a
# Ctrl+C / Ctrl+V flow and a "Copy as .bak" context-menu item; both
# land here. Files only (directories rejected with 400) and the new
# name is server-side derived so the API can't be tricked into
# clobbering anything: <stem><suffix>.bak, .bak.2, .bak.3 … picking
# the first non-existing name in the target directory.
# ============================================================
class CopyBakReq(BaseModel):
    src: str
    # Where to drop the .bak. Empty / omitted = same directory as src
    # (covers the "Ctrl+D / context-menu duplicate" path). Frontend
    # passes the currently-selected directory for cross-dir paste.
    dst_dir: str = ""


def _next_bak_name(parent: Path, original_name: str) -> str:
    """Pick the first non-existing <original_name>.bak[.N] under parent.

    Always appends `.bak`, even if original already ends in `.bak` (so
    `foo.txt.bak` → `foo.txt.bak.bak`). Increments via `.bak.2`, `.bak.3`,
    … This keeps the rule mechanical and predictable instead of trying
    to be clever about "already a backup".
    """
    base = f"{original_name}.bak"
    if not (parent / base).exists():
        return base
    # .bak exists → try .bak.2, .bak.3, … Cap at a sane upper bound so a
    # pathological directory full of .bak.N siblings can't hang the call.
    for i in range(2, 1000):
        cand = f"{original_name}.bak.{i}"
        if not (parent / cand).exists():
            return cand
    raise HTTPException(status_code=409, detail="too many .bak siblings")


@router.post("/copy-bak", dependencies=[Depends(require_token)])
def copy_bak(req: CopyBakReq) -> dict:
    src = safe_resolve(req.src)
    _guard_not_trash(src)
    if not src.exists():
        raise HTTPException(status_code=404, detail="source not found")
    # Files only. Directory copy is a different beast (shutil.copytree,
    # permission edge cases, can be slow on big trees) — out of scope for
    # the .bak shortcut.
    if src.is_dir():
        raise HTTPException(status_code=400, detail="directories not supported")
    if req.dst_dir:
        parent = safe_resolve(req.dst_dir)
        if not parent.exists() or not parent.is_dir():
            raise HTTPException(status_code=404, detail="dst_dir not found")
    else:
        parent = src.parent
    new_name = _next_bak_name(parent, src.name)
    dst = parent / new_name
    _guard_not_trash(dst)
    # safe_resolve the final path so the anti-traversal guard fires for
    # the destination too.
    dst_rel = str(dst.relative_to(ROOT))
    safe_resolve(dst_rel, allow_sensitive=True)
    # The appended `.bak[.N]` suffix means `_is_sensitive` (exact-name /
    # suffix match) never fires on the destination — `secrets.env.bak`
    # wouldn't match `.env.*`. Strip the trailing `.bak[.N]` chain and
    # re-check the underlying name so a copy can't smuggle a sensitive
    # file into a `.bak` wrapper. (Source is already blocked by the
    # default safe_resolve above; this hardens the dst symmetrically.)
    underlying = new_name
    while True:
        m = re.match(r"^(.*)\.bak(?:\.\d+)?$", underlying)
        if not m:
            break
        underlying = m.group(1)
    if _is_sensitive(Path(underlying)):
        raise HTTPException(status_code=403, detail="sensitive file blocked")
    shutil.copy2(src, dst)
    return {"ok": True, "path": dst_rel, "name": new_name}


SEARCH_IGNORE = {".git", "node_modules", "__pycache__", ".venv", "venv",
                 ".cache", ".pytest_cache", ".mypy_cache", "dist", "build",
                 # Trash always excluded from search/grep regardless of
                 # show_hidden — otherwise a search for "foo" surfaces every
                 # version of foo.md the user has ever deleted, which the
                 # trash UI is purpose-built to present separately.
                 TRASH_DIR_NAME}

GREP_EXTS = {".md", ".markdown", ".txt", ".html", ".htm", ".json", ".yaml", ".yml",
             ".py", ".js", ".ts", ".css", ".sh", ".toml", ".ini", ".csv", ".sql",
             ".log", ".xml", ".rst", ".tex"}


MAX_GREP_FILE_SIZE = 1_000_000   # 1MB per file — skip large files
MAX_GREP_TIME_SEC = 8            # soft time budget


# ============================================================
# Directory-listing cache for search / grep (2026-05-28).
#
# Both endpoints used to call `os.walk(ROOT)` from scratch on every
# request — for a 3000-file archive this is ~200-500 ms of pure
# scandir + per-entry stat() overhead even before any file content is
# touched. Most directories don't change between calls, so we cache
# `{dir_path: (mtime, [(name, is_dir), ...])}` keyed by directory
# mtime; a hit skips the scandir entirely. Filesystem mtime semantics
# guarantee a directory's mtime updates iff its entry list changes
# (add / remove / rename), which is exactly the cache invalidation
# trigger we need. File CONTENT changes do NOT bump parent mtime on
# ext4 / btrfs, but we deliberately don't cache file size / mtime —
# callers that need those `stat()` per file independently (fast).
#
# Thread-safe because FastAPI runs sync route handlers in a thread
# pool; two concurrent /api/files/search calls would race on the dict
# without _DIR_CACHE_LOCK.
# ============================================================
_DIR_CACHE: dict[str, tuple[float, list[tuple[str, bool]]]] = {}
_DIR_CACHE_LOCK = threading.Lock()
# Bound the cache so a pathological archive (millions of dirs) can't
# OOM the process. Typical personal archives have 50-500 dirs total,
# so this rarely matters. On overflow we drop the oldest insertion.
_DIR_CACHE_MAX = 5000


def _cached_walk(root: Path, ignore: set[str], show_hidden: bool):
    """Generator that mimics `os.walk(root)` but caches each directory's
    entry list by mtime, and applies the `ignore` / `show_hidden`
    filters in one pass.

    Yields `(dirpath: Path, dirnames: list[str], filenames: list[str])`.
    Callers that previously mutated `dirnames[:]` to filter no longer
    need to — this function pre-filters."""
    # Explicit stack so we control descent order + can interleave the
    # cache hit/miss path cleanly. DFS by `.pop()` matches os.walk's
    # top-down behavior for callers that bail on a time budget.
    stack: list[Path] = [root]
    while stack:
        dp = stack.pop()
        try:
            dir_mtime = dp.stat().st_mtime
        except OSError:
            continue
        key = str(dp)
        with _DIR_CACHE_LOCK:
            cached = _DIR_CACHE.get(key)
        entries: list[tuple[str, bool]]
        if cached is not None and cached[0] == dir_mtime:
            entries = cached[1]
        else:
            entries = []
            try:
                with os.scandir(dp) as it:
                    for de in it:
                        try:
                            entries.append(
                                (de.name, de.is_dir(follow_symlinks=False)))
                        except OSError:
                            continue
            except OSError:
                continue
            with _DIR_CACHE_LOCK:
                # Bound the cache via FIFO eviction (insertion-ordered
                # dict). Not strict LRU but cheap and good enough.
                if len(_DIR_CACHE) > _DIR_CACHE_MAX:
                    try:
                        _DIR_CACHE.pop(next(iter(_DIR_CACHE)))
                    except StopIteration:
                        pass
                _DIR_CACHE[key] = (dir_mtime, entries)
        dirnames: list[str] = []
        filenames: list[str] = []
        for name, is_dir in entries:
            if name in ignore:
                continue
            if not show_hidden and name.startswith("."):
                continue
            if is_dir:
                dirnames.append(name)
                stack.append(dp / name)
            else:
                filenames.append(name)
        yield dp, dirnames, filenames


# Concurrency gate: each grep can burn up to MAX_GREP_TIME_SEC of CPU while
# holding a threadpool thread. Without a cap, a few rapid keystrokes (or two
# devices searching at once) stack full-archive scans and starve the pool —
# every other endpoint (chat, sessions) stalls behind them. Two concurrent
# scans is plenty for interactive use; excess requests fail fast with 429
# rather than queueing (the UI debounces and just issues a fresh search).
_GREP_GATE = threading.BoundedSemaphore(2)


@router.get("/grep", dependencies=[Depends(require_token)])
def grep(q: str, limit: int = 50, show_hidden: bool = False) -> dict:
    """Cross-platform full-text search (pure Python, no grep dependency).
    Uses `_cached_walk` so the directory-listing phase is O(changed-dirs)
    instead of O(all-dirs) — repeat searches on a quiet archive only stat
    file contents, not the directory structure itself."""
    if not _GREP_GATE.acquire(blocking=False):
        raise HTTPException(429, "search busy — try again")
    try:
        return _grep_impl(q, limit, show_hidden)
    finally:
        _GREP_GATE.release()


def _grep_impl(q: str, limit: int, show_hidden: bool) -> dict:
    q_lower = q.strip().lower()
    # Minimum query length: a single character matches nearly every file
    # and always runs the full archive scan to the 8s time budget while
    # holding a threadpool thread hostage. Short queries early-return empty
    # (the UI debounces and only the user typing 1 char hits this).
    if len(q_lower) < 2:
        return {"hits": []}
    hits: list[dict] = []
    started = time.monotonic()
    timed_out = False
    # ROOT.resolve() is a loop invariant — hoist it out of the per-file loop
    # so the symlink-escape guard doesn't re-resolve ROOT once per candidate
    # file (was N stat-resolves per search).
    root_real = ROOT.resolve()
    for dirpath, _dirnames, filenames in _cached_walk(
            ROOT, SEARCH_IGNORE, show_hidden):
        if time.monotonic() - started > MAX_GREP_TIME_SEC:
            timed_out = True
            break
        for fname in filenames:
            # 隐藏文件即使没扩展名也允许 grep（用户主动开了 show_hidden 说明想看）
            if Path(fname).suffix.lower() not in GREP_EXTS and not (show_hidden and fname.startswith(".")):
                continue
            full = Path(dirpath) / fname
            try:
                # Symlink escape guard: resolve() follows symlinks, so a file
                # named `notes.md` inside ROOT pointing at /etc/passwd would
                # otherwise be opened and grepped. Confirm the real target is
                # still under ROOT before reading. Also run _is_sensitive on the
                # RESOLVED path so a symlink masking a `.env`/`*.pem` target is
                # caught (name-only check misses that).
                try:
                    resolved = full.resolve()
                except (OSError, ValueError):
                    continue
                if root_real != resolved and root_real not in resolved.parents:
                    continue
                if _is_sensitive(full) or _is_sensitive(resolved):
                    continue
                # File-level stat IS NOT cached (file content changes
                # don't bump parent dir mtime on ext4/btrfs, so a cached
                # size would lie). One stat per candidate is sub-µs.
                if full.stat().st_size > MAX_GREP_FILE_SIZE:
                    continue
                with full.open("r", encoding="utf-8", errors="replace") as f:
                    for i, line in enumerate(f, 1):
                        if q_lower in line.lower():
                            try:
                                rel = str(full.relative_to(ROOT))
                            except ValueError:
                                continue
                            hits.append({
                                "path": rel,
                                "name": fname,
                                "line": i,
                                "snippet": line.strip()[:200],
                            })
                            if len(hits) >= limit:
                                return {"hits": hits, "truncated": True}
            except OSError:
                continue
            if time.monotonic() - started > MAX_GREP_TIME_SEC:
                timed_out = True
                break
        if timed_out:
            break
    return {"hits": hits, "truncated": timed_out}


@router.get("/search", dependencies=[Depends(require_token)])
def search(q: str, limit: int = 100, show_hidden: bool = False) -> dict:
    """Filename / dirname substring search. Same `_cached_walk` win as
    grep — bigger here in relative terms because search only reads
    names (no file content), so the directory-listing IS the entire
    cost. Repeat searches over a quiet archive drop from ~200 ms to
    ~20 ms on a 3000-file tree."""
    q_lower = q.strip().lower()
    if not q_lower:
        return {"entries": []}
    hits: list[dict] = []
    for dirpath, dirnames, filenames in _cached_walk(
            ROOT, SEARCH_IGNORE, show_hidden):
        for name in dirnames + filenames:
            if q_lower in name.lower():
                full = Path(dirpath) / name
                try:
                    stat = full.stat()
                except OSError:
                    continue
                # is_dir() was called twice per hit (once for the flag, once
                # for the size branch) — each is a syscall. Compute once.
                is_dir = full.is_dir()
                hits.append({
                    "name": name,
                    "path": str(full.relative_to(ROOT)),
                    "is_dir": is_dir,
                    "size": stat.st_size if not is_dir else 0,
                    "mtime": stat.st_mtime,
                })
                if len(hits) >= limit:
                    return {"entries": hits, "truncated": True}
    return {"entries": hits, "truncated": False}
